from celery import shared_task
from sms.models import TaskHistory
from celery.task.schedules import crontab
from celery.decorators import periodic_task
from celery.utils.log import get_task_logger
from datetime import datetime
from collections import OrderedDict
from student.models import Subject, StudentClass, Course, Teach
from staff.models import Staff
from openpyxl import load_workbook
from . models import UploadedFile
from school.models import School

from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

import io
from django.core.files.uploadedfile import InMemoryUploadedFile
from tempfile import NamedTemporaryFile
logger = get_task_logger(__name__)

# A periodic task that will run every minute (the symbol "*" means every)
# @periodic_task(run_every=(crontab(hour="*", minute="*", day_of_week="*")), 
#     ignore_result=True)

def upload_file(file, file_name, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'):
    with open(file, 'rb') as f:
        file = f.read()
        name = "".join(file_name.split('.')[:len(file_name.split('.'))-1])
        file_data = InMemoryUploadedFile(
                file=io.BytesIO(file),
                field_name=file_name,
                name=file_name,
                content_type=content_type,
                size=len(file),
                charset='utf-8',
            )
        new_file = UploadedFile.objects.get_or_create(name=name)[0]
        new_file.file = file_data
        new_file.save()
        return new_file

# @shared_task
def generate_subject_teacher_template():
    school = School.objects.all().first()
    file = 'resources/assets/sheets/teacher_subjects.xlsx'

    try: 
        workbook = load_workbook(filename = file) 
        worksheet = workbook["teacher_subjects"]

        start_row = 15
        worksheet.cell(column = 1, row = start_row, value='SUBJECT NAME IN FULL')
        worksheet.cell(column = 2, row = start_row, value='CLASS NAME IN FULL')
        worksheet.cell(column = 3, row = start_row, value='STAFF ID')

        start_row += 1
        subjects = Subject.objects.all()
        for subject in subjects:
            if not subject.is_elective:
                classes  = StudentClass.objects.all()
                for clazz in classes:
                    worksheet.cell(column = 1, row = start_row, value=subject.name)
                    worksheet.cell(column = 2, row = start_row, value=clazz.name)            

                    combination  = Teach.objects.filter(subject=subject, student_class=clazz).first()
                    if combination:
                        if combination.staff:
                            worksheet.cell(column = 3, row = start_row, value=combination.staff.staff_id)
                    start_row += 1
            else:
                courses = Course.objects.all()
                for course in courses:
                    if subject in course.subjects.all():
                        classes = course.student_classes.all()
                        for clazz in classes:
                            worksheet.cell(column = 1, row = start_row, value=subject.name)
                            worksheet.cell(column = 2, row = start_row, value=clazz.name)
                            
                            combination  = Teach.objects.filter(subject=subject, student_class=clazz).first()
                            if combination:
                                if combination.staff:
                                    worksheet.cell(column = 3, row = start_row, value=combination.staff.staff_id)
                            start_row += 1
        
        new_file = "resources/assets/sheets/temp/teacher_subjects.xlsx"
        workbook.save(new_file)

        upload_file(new_file, "teacher_subjects.xlsx", 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as err:
        return str(err)
    return True
generate_subject_teacher_template()