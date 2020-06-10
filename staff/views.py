from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from dashboard.models import (RecordFile, 
                            UploadedFile, 
                            HouseMasterRequest, 
                            ClassTeacherRequest)
from dashboard.utils import (add_record,
                        add_house_master_remarks, 
                        add_class_teacher_remarks)
from student.models import (ClassTeacherRemarks, 
                            HouseMasterRemarks,
                            Course, Teach,
                            Student, StudentClass, Record, Subject)
from accounts.decorators import logged_in_as, login_required
from django.utils import timezone
from school.models import School
from helpers.functions import redirect_handler
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from dashboard.tasks import upload_file


@login_required
def welcome(request):    
    available_subjects = Subject.objects.filter()
    
    context = {
        "available_subjects":available_subjects,
    }
    return render(request, "staff/welcome.html", context)

@login_required
def select_class(request, subject_id):    
    subject = get_object_or_404(Subject, subject_id=subject_id)

    if request.method == "GET":
        if subject.is_elective:
            # Courses offering the subject
            courses = subject.courses.all()
            available_classes = StudentClass.objects.filter(course__in=courses)
        else:
            available_classes = StudentClass.objects.all()
        
        error_message = request.session.pop("error_message", None)
        context = {
            "subject":subject,
            "available_classes":available_classes,
            "error_message":error_message,
        }
        return render(request, "staff/class_selection.html", context)
    elif request.method == "POST":
        student_class_id = request.POST.get("student_class_id")
        student_class = get_object_or_404(StudentClass, student_class_id=student_class_id)
        
        teach = Teach.objects.get_or_create(subject = subject,
                                    student_class = student_class,
                                    staff=request.user.staff)
        if not teach[1]:
            request.session["error_message"] = f"{subject.name} is being taught in {student_class.name} by {teach[0].staff.name}."
            return redirect("staff:select_class", subject.subject_id)
        return redirect("staff:input_subject_records")

# Drop subjects
@logged_in_as(user_type="teacher")
def drop_subjects(request):    
    if request.method == "GET":
        teaches = request.user.staff.teaches.all()
        context = {
            "teaches":teaches,
        }
        return render(request, "staff/drop_subjects.html", context)

    elif request.method == "POST":
        teach_id = request.POST.get("teach_id")
        teach = Teach.objects.filter(id=teach_id).delete()
        return redirect("staff:welcome")


# Subject teacher records
@logged_in_as(user_type="teacher")
def subject_teacher(request):
    semester = request.school.current_semester
    academic_year = request.school.current_academic_year

    teaches = request.user.staff.teaches.all()
    pending_ids = []
    submitted_ids = []
    for teach in teaches:
        record = Record.objects.filter( 
                                academic_year=academic_year,
                                semester = semester,
                                subject = teach.subject,
                                record_class = teach.student_class
                            )
        if not record:
            pending_ids.append(teach.id)
        else:
            submitted_ids.append(teach.id)

    pending = Teach.objects.filter(id__in=pending_ids)
    submitted = Teach.objects.filter(id__in=submitted_ids)
    
   # Retrieve upload_record message from user session
    message = request.session.pop("message", None)
    error_message = request.session.pop("error_message", None)
    context = {
        "pending": pending,
        "submitted": submitted,
        "message": message,
        "error_message":error_message,
    }
    return render(request, "staff/records.html", context)

@logged_in_as(user_type="formteacher")
def form_teacher(request):
    semester = request.school.current_semester
    academic_year = request.school.current_academic_year
    klass = request.user.staff.student_class

    remarks = ClassTeacherRemarks.objects.filter(
                semester = semester,
                academic_year  = academic_year,
                student_class = klass
            )
    warnings = ''
    print(remarks.count() == klass.students.all().count())
    if remarks:
        if not remarks.count() == klass.students.all().count():
            warnings = f"Could not find remark for some students  in {klass}"

    # Retrieve upload_record message from user session
    message = request.session.pop("message", None)
    error_message = request.session.pop("error_message", None)
    context = {
        "remarks": remarks,
        "warnings": warnings,
        "message": message,
        "error_message":error_message,
    }
    return render(request, "staff/form_teacher.html", context)

@logged_in_as(user_type="housemaster")
def house_master(request):
    semester = request.school.current_semester
    academic_year = request.school.current_academic_year

    # Students of the house
    house = request.user.staff.housemaster.house
    form_1_students  = Student.objects.filter(house=house, student_class__form = 1)
    form_2_students  = Student.objects.filter(house=house, student_class__form = 2)
    form_3_students  = Student.objects.filter(house=house, student_class__form = 3)
    
    form_1_remark = HouseMasterRemarks.objects.filter(
                semester = semester,
                academic_year  = academic_year,
                student__house = house,
                student__student_class__form = 1
            )
    form_2_remark = HouseMasterRemarks.objects.filter(
                semester = semester,
                academic_year  = academic_year,
                student__house = house,
                student__student_class__form = 2
            )

    form_3_remark = HouseMasterRemarks.objects.filter(
        semester = semester,
        academic_year  = academic_year,
        student__house = house,
        student__student_class__form = 3
    )

    warnings = []
    if form_1_remark:
        if not len(form_1_remark) == len(form_1_students):
            warnings.append(f"Could not find remark for some form 1 students  in {house} house")
    
    if form_2_remark:
        if not len(form_2_remark) == len(form_2_students):
            warnings.append(f"Could not find remark for some form 2 students  in {house} house")
    
    if form_3_remark:
        if not len(form_3_remark) == len(form_3_students):
             warnings.append(f"Could not find remark for some form 3 students  in {house} house")
    
    # Retrieve upload_record message from user session
    message = request.session.pop("message", None)
    error_message = request.session.pop("error_message", None)
    context = {
        "form_1_remark":form_1_remark,
        "form_2_remark":form_2_remark,
        "form_3_remark":form_3_remark,
        "warnings":warnings,
        "message": message,
        "error_message":error_message,
    }
    return render(request, "staff/house_master.html", context)

@logged_in_as(user_type="formteacher")
def student_reports(request):
    academic_years   = Record.objects.all().values("academic_year").distinct()
    semesters        = Record.objects.all().values("semester").distinct()
    students         = Student.objects.filter(student_class=request.user.staff.student_class)

    context = {
        "academic_years": academic_years,
        "students": students,
        "semesters": semesters,
        "classes": StudentClass.objects.filter(name=request.user.staff.student_class.name),
    }
    return render(request, "staff/student_reports.html", context)

# Subject teacher upload functions
@logged_in_as(user_type="teacher")
def upload_record(request):
    error_message = ""
    if request.method == "POST":
        excel_file = request.FILES.get("file")
        
        academic_year   = request.POST.get("academic_year")
        form            = request.POST.get("form")
        semester        = request.POST.get("semester")
        subject_name    = request.POST.get("subject_name")

        # Parameters to the redirect function
        options = {
            "academic_year":academic_year,
            "form":form,
            "semester":semester,
            "subject_name":subject_name,
        }

        file_extention = str(excel_file).split(".")[-1]
        if (file_extention == "xls") or file_extention == "xlsx":
            if add_record(request, excel_file):
                # User uploaded file
                uploadedfile = UploadedFile(name=excel_file.name, file=excel_file, user=request.user)
                uploadedfile.save()
                return redirect_handler(request, "staff:preview_added_records", options)
                # return redirect()
        else:
            request.session["error_message"] = "Only excel files are acceptable!"
    return redirect("staff:preview_added_records")

# Form teacher upload functions
@logged_in_as(user_type="formteacher")
def upload_form_teacher_remarks(request):
    error_message = ""
    if request.method == "POST":
        excel_file = request.FILES.get("file")
        # request_id = request.POST.get("request_id")
        
        file_extention = str(excel_file).split(".")[-1]
        if (file_extention == "xls") or file_extention == "xlsx":
            # Add Record
            if add_class_teacher_remarks(request, excel_file):
                # User uploaded file
                uploadedfile = UploadedFile(name=excel_file.name, file=excel_file, user=request.user)
            return redirect("staff:form_teacher")
        else:
            request.session["error_message"] = "Only excel files are acceptable!"
    return redirect("staff:form_teacher")

# House master upload functions
@logged_in_as(user_type="housemaster")
def upload_house_master_remarks(request):
    error_message = ""
    if request.method == "POST":
        excel_file = request.FILES.get("file")
        
        file_extention = str(excel_file).split(".")[-1]
        if (file_extention == "xls") or file_extention == "xlsx":
            # Add Record
            if add_house_master_remarks(request, excel_file):
                 # User uploaded file
                uploadedfile = UploadedFile(name=excel_file.name, file=excel_file, user=request.user)
                uploadedfile.save()        
            return redirect("staff:house_master")
        else:
            request.session["error_message"] = "Only excel files are acceptable!"
    return redirect("staff:house_master")

@logged_in_as(user_type="formteacher")
def input_form_master_remarks(request):
    context = {
        "school":School.objects.all().first()
    }
    return render(request, "staff/input_form_master_remarks.html", context)


@logged_in_as(user_type="teacher")
def input_subject_records(request):
    classes = request.user.staff.teaches.all().values("student_class__name").distinct()
    subjects = request.user.staff.teaches.all().values("subject__name").distinct()

    if not subjects: return redirect("staff:welcome")
    
    context = {
        "school":School.objects.all().first(),
        "classes":classes,
        "subjects":subjects,
    }
    return render(request, "staff/input_subject_records.html", context)

@logged_in_as(user_type="housemaster")
def input_house_master_remarks(request):
    
    context = {
     
    }
    return render(request, "staff/input_house_master_remarks.html", context)

def preview_added_records(request):
    academic_year   =   request.GET.get("academic_year")
    semester        =   request.GET.get("semester")
    form            =   request.GET.get("form")
    subject_name    =   request.GET.get("subject_name")

    subject = Subject.objects.filter(name=subject_name).first()
    if subject.is_elective:
        students = subject.students.filter(
                        student_class__form = form,
                        student_class__in=request.user.staff.teaches.filter(subject=subject).values("student_class"))
    else:
        students  = Student.objects.filter(
                        student_class__form = form,
                        student_class__in=request.user.staff.teaches.filter(subject=subject).values("student_class"))
    

    s = "".join(semester.split(" "))
    sbj = "".join(subject_name.split(" "))
    ay = "_".join(academic_year.split("/"))
    file_name       = f"{ay}_{s}_{sbj}_form_{form}"
    file            = f"resources/assets/sheets/academic_records.xlsx"

    workbook = load_workbook(filename = file) 
    worksheet = workbook["records"]

    worksheet.cell(column = 1, row = 15, value="No. on Roll")
    worksheet.cell(column = 1, row = 16, value="Academic Year")
    worksheet.cell(column = 1, row = 17, value="Form")
    worksheet.cell(column = 1, row = 18, value="Subject Name")
    worksheet.cell(column = 1, row = 19, value="Semester")

    worksheet.cell(column = 2, row = 15, value=students.count())
    worksheet.cell(column = 2, row = 16, value=academic_year)
    worksheet.cell(column = 2, row = 17, value=form)
    worksheet.cell(column = 2, row = 18, value=subject_name)
    worksheet.cell(column = 2, row = 19, value=semester)

    start_row = 21
    worksheet.cell(column = 1, row = start_row, value='STUDENT ID')
    worksheet.cell(column = 2, row = start_row, value='FULL NAME')
    worksheet.cell(column = 3, row = start_row, value='CLASS SCORE')
    worksheet.cell(column = 4, row = start_row, value='EXAM SCORE')
    worksheet.cell(column = 5, row = start_row, value='TOTAL')
    worksheet.cell(column = 6, row = start_row, value='GRADE')
    worksheet.cell(column = 7, row = start_row, value='REMARK')
    worksheet.cell(column = 8, row = start_row, value='POSITION')

    start_row += 1
    students_records = {}
    for student in students.order_by("student_id"):
        record = Record.objects.filter(
                    student = student,
                    subject__name = subject_name,
                    academic_year=academic_year, 
                    semester=semester).order_by("-id").first()
        students_records.update({student:record})

        worksheet.cell(column = 1, row = start_row, value=student.student_id)
        worksheet.cell(column = 2, row = start_row, value=student.name)

        if record:
            worksheet.cell(column = 3, row = start_row, value=record.class_score)
            worksheet.cell(column = 4, row = start_row, value=record.exam_score)
            worksheet.cell(column = 5, row = start_row, value=record.total)
            worksheet.cell(column = 6, row = start_row, value=record.grade)
            worksheet.cell(column = 7, row = start_row, value=record.remark)
            worksheet.cell(column = 8, row = start_row, value=record.position)
        else:
            worksheet.cell(column = 3, row = start_row, value="")
            worksheet.cell(column = 4, row = start_row, value="")
            worksheet.cell(column = 5, row = start_row, value="")
            worksheet.cell(column = 6, row = start_row, value="")
            worksheet.cell(column = 7, row = start_row, value="")
            worksheet.cell(column = 8, row = start_row, value="")

        start_row += 1

    try:
        s = "".join(semester.split(" "))
        sbj = "".join(subject_name.split(" "))
        ay = "_".join(academic_year.split("/"))
        file_name       = f"{ay}_{s}_{sbj}_form_{form}_filled"

        new_file = "resources/assets/sheets/temp/academic_records_filled.xlsx"
        workbook.save(new_file)
        sheet = upload_file(new_file, file_name+".xlsx")
    
    except Exception as err:
        return HttpResponse(str(err))
    
    message = request.session.pop("message", None)
    context = {
        "students_records":students_records,
        "sheet":sheet,
        "students":students,
        "message":message,
    }
    return render(request, "staff/preview_added_records.html", context)