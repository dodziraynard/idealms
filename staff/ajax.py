from student.models import (Student, 
                            ClassTeacherRemarks, 
                            HouseMasterRemarks,
                            StudentClass, Record,
                            Subject)
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from accounts.decorators import login_required
from dashboard.tasks import upload_file
from dashboard.models import UploadedFile
from openpyxl.workbook import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

@login_required
def get_form_teacher_remark_input_template(request):
    if not request.method == "POST":
        return HttpResponse("Noops!")
    academic_year   =   request.POST.get("academic_year")
    semester    =   request.POST.get("semester")
    class_name  =   request.POST.get("class_name")
    
    students = Student.objects.filter(student_class__name = class_name)

    students_remark = {}
    total_attendance = ""
    for student in students:
        remark = ClassTeacherRemarks.objects.filter(
                    student = student,
                    academic_year=academic_year, 
                    student_class__name=class_name,
                    semester=semester).first()
        students_remark.update({student:remark})
        if remark:
            total_attendance = remark.total_attendance

    context = {
        "students_remark":students_remark,
        "academic_year":academic_year,
        "semester":semester,
        "class_name":class_name,
        "total_attendance":total_attendance,
    }
    return render(request, "staff/ajax/form_teacher_remark_template.html", context)
    
@login_required
def save_form_teacher_remark(request):
    if request.method == "POST":
        total_attendance    = request.POST.get("total_attendance")
        attendance          = request.POST.get("attendance")
        attitude            = request.POST.get("attitude")
        conduct             = request.POST.get("conduct")
        interest            = request.POST.get("interest")
        remark              = request.POST.get("remark")
        student_id          = request.POST.get("student_id")
        semester            = request.POST.get("semester")
        academic_year       = request.POST.get("academic_year")
        class_name          = request.POST.get("class_name")

        student = Student.objects.filter(student_id=student_id).first()
        student_class = StudentClass.objects.filter(name=class_name).first()
        if student and student_class:
            new_remark = ClassTeacherRemarks.objects.get_or_create(
                                    academic_year=academic_year, 
                                    student = student,
                                    student_class= student_class,
                                    semester=semester)[0]
            
            new_remark.total_attendance = total_attendance
            new_remark.attendance = attendance
            new_remark.attitude = attitude
            new_remark.conduct = conduct
            new_remark.interest = interest
            new_remark.remark = remark
            new_remark.save()
            return HttpResponse("Saved successfully")

# Last class teacher remark
@login_required
def get_previous_remark(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        option = request.POST.get("option")

        remark = ClassTeacherRemarks.objects.filter(student__student_id=student_id).order_by("-id").first()
        if remark:
            if option == "attitude":
                return HttpResponse(remark.attitude)
            if option == "conduct":
                return HttpResponse(remark.conduct)
            if option == "interest":
                return HttpResponse(remark.interest)
            if option == "remark":
                return HttpResponse(remark.remark)
    return HttpResponse(None)

@login_required
def get_academic_record(request):
    student_id          = request.POST.get("student_id")
    semester            = request.POST.get("semester")
    academic_year       = request.POST.get("academic_year")
    class_name          = request.POST.get("class_name")
    
    student        = Student.objects.filter(student_id=student_id).first()
    record_class   = StudentClass.objects.filter(name=class_name).first()

    records = Record.objects.filter(student=student,
                                    semester = semester,
                                    academic_year = academic_year,
                                    record_class = record_class
                        )
    context = {
        "records":records,
        "student":student,
        "record_class":record_class,
        "semester":semester,
        "academic_year":academic_year,
    }
    return render(request, "staff/ajax/terminal_report.html", context)

@login_required
def get_subject_score_input_template(request):
    if not request.method == "POST":
        return HttpResponse("Nops!")

    academic_year   =   request.POST.get("academic_year")
    semester        =   request.POST.get("semester")
    form            =   request.POST.get("form")
    subject_name    =   request.POST.get("subject_name")
    
    subject = Subject.objects.filter(name=subject_name).first()
    if subject.is_elective:
        students = subject.students.filter(
                        student_class__form = form,
                        student_class__in=request.user.staff.teaches.filter(subject=subject).values("student_class"))
    else:
        students  = Student.objects.filter(
                        student_class__form = form,
                        student_class__in=request.user.staff.teaches.filter(subject=subject).values("student_class"))
    
    students_records = {}
    for student in students.order_by("student_id"):
        records = Record.objects.filter(
                    student = student,
                    subject__name = subject_name,
                    academic_year=academic_year, 
                    semester=semester).order_by("-id").first()
        students_records.update({student:records})

    context = {
        "students_records":students_records,
        "academic_year":academic_year,
        "semester":semester,
        "subject_name":subject_name,
        "students":students,
    }
    return render(request, "staff/ajax/subject_record_template.html", context)

@login_required
def save_subject_record(request):
    if request.method == "POST":
        semester = request.POST.get("semester")
        student_id = request.POST.get("student_id")
        subject_name = request.POST.get("subject_name")
        academic_year = request.POST.get("academic_year")
        class_score = request.POST.get("class_score")
        exam_score = request.POST.get("exam_score")
        total = request.POST.get("total")
        grade = request.POST.get("grade")
        remark = request.POST.get("remark")
        position = request.POST.get("position")
        roll_no = request.POST.get("roll_no")

        student = Student.objects.filter(student_id=student_id).first()
        subject = Subject.objects.filter(name=subject_name).first()
        new_record = Record.objects.get_or_create(
            semester = semester,
            student = student,
            academic_year = academic_year,
            subject = subject
        )[0]

        new_record.class_score  = class_score
        new_record.exam_score   = exam_score
        new_record.position     = position
        new_record.roll_no      = roll_no
        new_record.save()
        return HttpResponse("Saved successfully")

# House master remark input template
@login_required
def get_house_master_input_template(request):
    if not request.method == "POST":
        return HttpResponse("Nops!")

    academic_year   =   request.POST.get("academic_year")
    semester        =   request.POST.get("semester")
    form            =   request.POST.get("form")
    
    students = Student.objects.filter(house=request.user.staff.housemaster.house.title(), student_class__form=form)
    students_remark = {}
    for student in students.order_by("student_id"):
        remark = HouseMasterRemarks.objects.filter(
                    student = student,
                    academic_year=academic_year, 
                    semester=semester).order_by('-id').first()
        students_remark.update({student:remark})
    context = {
        "students_remark":students_remark,
        "academic_year":academic_year,
        "semester":semester,
        "students":students,
    }
    return render(request, "staff/ajax/house_master_remark_template.html", context)

@login_required
def save_house_master_remark(request):
    if request.method == "POST":
        remark              = request.POST.get("remark")
        student_id          = request.POST.get("student_id")
        semester            = request.POST.get("semester")
        academic_year       = request.POST.get("academic_year")
        
        student = Student.objects.filter(student_id=student_id).first()
        if student:
            new_remark = HouseMasterRemarks.objects.get_or_create(
                                    academic_year=academic_year, 
                                    student = student,
                                    student_class= student.student_class,
                                    semester=semester)[0]
            new_remark.remark = remark
            new_remark.save()
            return JsonResponse({"response":"Saved successfully"})
        return JsonResponse({"response":"No students found"})
    return JsonResponse({"response":"Nops"})

# Last house master remark
@login_required
def get_previous_house_master_remark(request):
    if request.method == "POST":
        student_id = request.POST.get("student_id")
        remark = HouseMasterRemarks.objects.filter(student__student_id=student_id).order_by("-id").first()
        if remark:
            return HttpResponse(remark.remark)
    return HttpResponse(None)


# Teacher generating students sheets
@login_required
def generate_record_sheet(request):
    if request.method == "GET":
        return HttpResponse("Bad Request", status=400)
    academic_year   =   request.POST.get("academic_year")
    semester        =   request.POST.get("semester")
    form            =   request.POST.get("form")
    subject_name    =   request.POST.get("subject_name")
    
    # What a teacher teaches
    teaches = request.user.staff.teaches
    subject = Subject.objects.filter(name=subject_name).first()
    if subject.is_elective:
        students = subject.students.filter(
                        student_class__form = form,
                        student_class__in=teaches.filter(subject=subject).values("student_class"))
    else:
        students  = Student.objects.filter(
                        student_class__form = form,
                        student_class__in=teaches.filter(subject=subject).values("student_class"))
    
    s = "".join(semester.split(" "))
    sbj = "".join(subject_name.split(" "))
    ay = "_".join(academic_year.split("/"))
    file_name       = f"{ay}_{s}_{sbj}_form_{form}"
    file            = f"resources/assets/sheets/academic_records.xlsx"

    workbook = load_workbook(filename = file) 
    worksheet = workbook["records"]
    
    worksheet.cell(column = 1, row = 15, value="No. on Roll")
    worksheet.cell(column = 1, row = 16, value="Academic Year")
    worksheet.cell(column = 1, row = 17, value="Form")
    worksheet.cell(column = 1, row = 18, value="Subject Name")
    worksheet.cell(column = 1, row = 19, value="Semester")

    worksheet.cell(column = 2, row = 15, value=students.count())
    worksheet.cell(column = 2, row = 16, value=academic_year)
    worksheet.cell(column = 2, row = 17, value=form)
    worksheet.cell(column = 2, row = 18, value=subject_name)
    worksheet.cell(column = 2, row = 19, value=semester)

    start_row = 21
    worksheet.cell(column = 1, row = start_row, value='STUDENT ID')
    worksheet.cell(column = 2, row = start_row, value='FULL NAME')
    worksheet.cell(column = 3, row = start_row, value='CLASS SCORE')
    worksheet.cell(column = 4, row = start_row, value='EXAM SCORE')
    
    start_row += 1
    students_records = {}
    for student in students.order_by("student_id"):
        record = Record.objects.filter(
                    student = student,
                    subject__name = subject_name,
                    academic_year=academic_year, 
                    semester=semester).order_by("-id").first()
        students_records.update({student:record})

        worksheet.cell(column = 1, row = start_row, value=student.student_id)
        worksheet.cell(column = 2, row = start_row, value=student.name)

        if record:
            worksheet.cell(column = 3, row = start_row, value=record.class_score)
            worksheet.cell(column = 4, row = start_row, value=record.exam_score)
        else:
            worksheet.cell(column = 3, row = start_row, value="")
            worksheet.cell(column = 4, row = start_row, value="")

        start_row += 1
    
    try:
        new_file = "resources/assets/sheets/temp/academic_records.xlsx"
        workbook.save(new_file)
        sheet = upload_file(new_file, file_name+".xlsx")
        filled_sheet = UploadedFile.objects.filter(name=file_name+"_filled").first()

        context = {
            "students_records":students_records,
            "academic_year":academic_year,
            "form":form,
            "semester":semester,
            "subject_name":subject_name,
            "students":students,
            "sheet":sheet,
            "filled_sheet":filled_sheet
        }
        return render(request, "staff/ajax/subject_record_generated_sheet.html", context)
    except Exception as err:
        return HttpResponse(str(err))

    context = {
            "sheet":None
        }
    return render(request, "staff/ajax/subject_record_generated_sheet.html", context)


# Generate excel sheet for house master
def generate_house_master_remark_sheet(request):
    if not request.method == "POST":
        return HttpResponse("Nops!")

    academic_year   =   request.POST.get("academic_year")
    semester        =   request.POST.get("semester")
    form            =   request.POST.get("form")
    
    house       = request.user.staff.housemaster.house.title()
    students    = Student.objects.filter(house=house, student_class__form=form)
    
    # Populating the excel file
    s               = "".join(semester.split(" "))
    ay              = "_".join(academic_year.split("/"))
    file_name       = f"{ay}_{s}_{house}_form_{form}"
    file            = f"resources/assets/sheets/house_master_remarks.xlsx"
    
    workbook = load_workbook(filename = file) 
    worksheet = workbook["house_master_remarks"]

    worksheet.cell(column = 1, row = 15, value="HOUSE NAME")
    worksheet.cell(column = 1, row = 16, value="ACADEMIC YEAR")
    worksheet.cell(column = 1, row = 17, value="SEMESTER")

    worksheet.cell(column = 2, row = 15, value=house)
    worksheet.cell(column = 2, row = 16, value=academic_year)
    worksheet.cell(column = 2, row = 17, value=semester)

    start_row = 18
    worksheet.cell(column = 1, row = start_row, value='STUDENT ID')
    worksheet.cell(column = 2, row = start_row, value='FULL NAME')
    worksheet.cell(column = 3, row = start_row, value='REAMRK')

    sheet = None
    start_row += 1
    students_remark = {}
    for student in students.order_by("student_id"):
        remark = HouseMasterRemarks.objects.filter(
                    student = student,
                    academic_year=academic_year, 
                    semester=semester).order_by('-id').first()
        students_remark.update({student:remark})
        
        worksheet.cell(column = 1, row = start_row, value=student.student_id)
        worksheet.cell(column = 2, row = start_row, value=student.name)

        if remark:
            worksheet.cell(column = 3, row = start_row, value=remark.remark)
        else:
            worksheet.cell(column = 3, row = start_row, value="")

        start_row += 1

    new_file = "resources/assets/sheets/temp/house_master_remarks.xlsx"
    try:
        workbook.save(new_file)
    except Exception as err:
        return HttpResponse(str(err))

    file_name = f"{academic_year}_{semester}_{house}_house_{form}"
    sheet = upload_file(new_file, file_name+".xlsx")

    context = {
        "sheet":sheet,
        "students_remark":students_remark,
    }
    return render(request, "staff/ajax/house_master_remark_generated_sheet.html", context)

def generate_form_master_remark_sheet(request):
    if not request.method == "POST":
        return HttpResponse("Noops!")
    academic_year   =   request.POST.get("academic_year")
    semester    =   request.POST.get("semester")
    class_name  =   request.POST.get("class_name")
    
    students = Student.objects.filter(student_class__name = class_name)
    clazz = StudentClass.objects.filter(name=class_name).first()

    # Populating the excel file
    s               = "".join(semester.split(" "))
    ay              = "_".join(academic_year.split("/"))
    file_name       = f"{ay}_{s}_{class_name}_remarks"
    file            = f"resources/assets/sheets/class_teacher_remarks.xlsx"
    
    workbook = load_workbook(filename = file) 
    worksheet = workbook["class_teacher_remarks"]

    worksheet.cell(column = 1, row = 15, value="CLASS NAME")
    worksheet.cell(column = 1, row = 16, value="ACADEMIC YEAR")
    worksheet.cell(column = 1, row = 17, value="SEMESTER")

    r = ClassTeacherRemarks.objects.filter(
                    student = students.first(),
                    academic_year=academic_year, 
                    student_class__name=class_name,
                    semester=semester).first()
    
    total_attendance = r.total_attendance if r else ""

    worksheet.cell(column = 1, row = 18, value="TOTAL ATENDANCE")

    worksheet.cell(column = 2, row = 15, value=clazz.name)
    worksheet.cell(column = 2, row = 16, value=academic_year)
    worksheet.cell(column = 2, row = 17, value=semester)
    worksheet.cell(column = 2, row = 18, value=total_attendance)

    start_row = 19
    worksheet.cell(column = 1, row = start_row, value='STUDENT ID')
    worksheet.cell(column = 2, row = start_row, value='FULL NAME')
    worksheet.cell(column = 3, row = start_row, value='ATTENDANCE')
    worksheet.cell(column = 4, row = start_row, value='ATTITUDE')
    worksheet.cell(column = 5, row = start_row, value='INTEREST')
    worksheet.cell(column = 6, row = start_row, value='CONDUCT')
    worksheet.cell(column = 7, row = start_row, value='REMARK')

    start_row += 1
    students_remark = {}
    for student in students:
        remark = ClassTeacherRemarks.objects.filter(
                    student = student,
                    academic_year=academic_year, 
                    student_class__name=class_name,
                    semester=semester).first()
        students_remark.update({student:remark})

        worksheet.cell(column = 1, row = start_row, value=student.student_id)
        worksheet.cell(column = 2, row = start_row, value=student.name.upper())
        
        if remark:
            worksheet.cell(column = 3, row = start_row, value=remark.attendance)
            worksheet.cell(column = 4, row = start_row, value=remark.attitude)
            worksheet.cell(column = 5, row = start_row, value=remark.interest)
            worksheet.cell(column = 6, row = start_row, value=remark.conduct)
            worksheet.cell(column = 7, row = start_row, value=remark.remark)
        
        start_row += 1
    
    new_file = "resources/assets/sheets/temp/class_teacher_remarks.xlsx"
    try:
        workbook.save(new_file)
    except Exception as err:
        return HttpResponse(str(err))    
    sheet = upload_file(new_file, file_name+".xlsx")

    context = {
        "sheet":sheet,
        "students_remark":students_remark,
        "total_attendance":total_attendance,
    }
    return render(request, "staff/ajax/form_master_remark_generated_sheet.html", context)